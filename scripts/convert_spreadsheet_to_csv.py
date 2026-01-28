#!/usr/bin/env python3
"""
Convert Excel (.xlsx) or OpenDocument (.ods) spreadsheet to CSV while preserving text formatting.
Text formatting (bold, italic, underline) is converted to markdown syntax.

Usage:
    python convert_spreadsheet_to_csv.py input_file.xlsx output_prompt.csv output_response.csv

Expected spreadsheet format:
- Column 1: Contains "prompt" (for black cards) or "response" (for white cards)
- Column 2: The card text (with optional formatting)
- Column 3: Optional additional text to append to column 2

The script processes each row and categorizes cards based on the label in column 1.
"""

import sys
import csv
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.cell.rich_text import TextBlock, CellRichText
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)

try:
    from odf import text, teletype
    from odf.opendocument import load as load_ods
    from odf.table import Table, TableRow, TableCell
    HAS_ODF = True
except ImportError:
    HAS_ODF = False
    print("Warning: odfpy not installed. ODS support disabled. Install with: pip install odfpy")


def smart_strip(text):
    """Strip leading/trailing whitespace while preserving internal newlines.

    This removes spaces/tabs from the start and end, and removes leading/trailing
    blank lines, but preserves newlines within the text.
    """
    if not text:
        return ""

    # Split into lines
    lines = text.split('\n')

    # Remove leading empty lines
    while lines and not lines[0].strip():
        lines.pop(0)

    # Remove trailing empty lines
    while lines and not lines[-1].strip():
        lines.pop()

    # Strip each line but preserve the newlines between them
    stripped_lines = [line.strip() for line in lines]

    return '\n'.join(stripped_lines)


def convert_rich_text_to_markdown(cell):
    """Convert Excel rich text formatting to markdown.

    Preserves newlines within the cell text.
    """
    if cell.value is None:
        return ""

    # If it's a CellRichText object (formatted text)
    if isinstance(cell.value, CellRichText):
        result = []
        for item in cell.value:
            if isinstance(item, TextBlock):
                text_content = item.text
                font = item.font

                if font:
                    # Apply markdown formatting based on font properties
                    if font.b:  # Bold
                        text_content = f"**{text_content}**"
                    if font.i:  # Italic
                        text_content = f"*{text_content}*"
                    if font.u:  # Underline (use HTML since markdown doesn't have underline)
                        text_content = f"<u>{text_content}</u>"

                result.append(text_content)
            else:
                # Plain text part - preserve as-is (including newlines)
                result.append(str(item))

        return ''.join(result)

    # Check if the entire cell has formatting
    # Convert to string but preserve newlines
    text_content = str(cell.value)
    if cell.font:
        if cell.font.b:
            text_content = f"**{text_content}**"
        if cell.font.i:
            text_content = f"*{text_content}*"
        if cell.font.u:
            text_content = f"<u>{text_content}</u>"

    return text_content


def process_xlsx(input_file, output_prompt, output_response):
    """Process Excel (.xlsx) file."""
    print(f"Loading Excel file: {input_file}")
    wb = load_workbook(input_file, rich_text=True)
    ws = wb.active

    # Get all rows
    rows = list(ws.iter_rows())
    if not rows:
        print("Error: Empty spreadsheet")
        return

    prompt_cards = []
    response_cards = []

    # Process all rows - first column contains "prompt" or "response" label
    # Second column (and third if exists) contain the card data
    for row_idx, row in enumerate(rows, start=1):
        if len(row) < 2:
            continue

        # Get the label from first column
        label_cell = row[0]
        label = convert_rich_text_to_markdown(label_cell).strip().lower()

        # Get card text from second column (preserve newlines)
        card_cell = row[1]
        card_text = smart_strip(convert_rich_text_to_markdown(card_cell))

        # Get extra text from third column if it exists (preserve newlines)
        extra_text = ""
        if len(row) > 2:
            extra_cell = row[2]
            extra_text = smart_strip(convert_rich_text_to_markdown(extra_cell))

        # Combine card text with extra text if present
        if card_text:
            # If extra text exists, combine with a space (or newline if card_text ends with one)
            if extra_text:
                if card_text.endswith('\n'):
                    full_text = f"{card_text}{extra_text}"
                else:
                    full_text = f"{card_text} {extra_text}"
            else:
                full_text = card_text

            # Categorize based on label
            if "prompt" in label:
                prompt_cards.append([full_text])
                # Show preview with newlines replaced for display
                preview = full_text.replace('\n', '\\n')[:50]
                print(f"Row {row_idx}: Black card - {preview}...")
            elif "response" in label:
                response_cards.append([full_text])
                preview = full_text.replace('\n', '\\n')[:50]
                print(f"Row {row_idx}: White card - {preview}...")
    
    # Write prompt cards CSV
    if prompt_cards:
        with open(output_prompt, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Card Text', 'Tag1', 'Tag2', 'Tag3', 'Tag4', 'Tag5', 'Tag6', 'Tag7', 'Tag8', 'Tag9', 'Tag10'])
            writer.writerows(prompt_cards)
        print(f"Wrote {len(prompt_cards)} prompt cards to {output_prompt}")
    
    # Write response cards CSV
    if response_cards:
        with open(output_response, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Card Text', 'Tag1', 'Tag2', 'Tag3', 'Tag4', 'Tag5', 'Tag6', 'Tag7', 'Tag8', 'Tag9', 'Tag10'])
            writer.writerows(response_cards)
        print(f"Wrote {len(response_cards)} response cards to {output_response}")


def process_ods(input_file, output_prompt, output_response):
    """Process OpenDocument (.ods) file."""
    if not HAS_ODF:
        print("Error: odfpy is required for ODS files. Install it with: pip install odfpy")
        sys.exit(1)

    print(f"Loading ODS file: {input_file}")
    doc = load_ods(input_file)

    # Get the first sheet
    sheets = doc.spreadsheet.getElementsByType(Table)
    if not sheets:
        print("Error: No sheets found in ODS file")
        return

    sheet = sheets[0]
    rows = sheet.getElementsByType(TableRow)

    if not rows:
        print("Error: Empty spreadsheet")
        return

    prompt_cards = []
    response_cards = []

    # Process all rows - first column contains "prompt" or "response" label
    # Second column (and third if exists) contain the card data
    for row_idx, row in enumerate(rows, start=1):
        cells = row.getElementsByType(TableCell)

        if len(cells) < 2:
            continue

        # Get the label from first column
        label = teletype.extractText(cells[0]).strip().lower()

        # Get card text from second column (preserve newlines)
        card_text = smart_strip(teletype.extractText(cells[1]))

        # Get extra text from third column if it exists (preserve newlines)
        extra_text = ""
        if len(cells) > 2:
            extra_text = smart_strip(teletype.extractText(cells[2]))

        # Combine card text with extra text if present
        if card_text:
            # If extra text exists, combine with a space (or newline if card_text ends with one)
            if extra_text:
                if card_text.endswith('\n'):
                    full_text = f"{card_text}{extra_text}"
                else:
                    full_text = f"{card_text} {extra_text}"
            else:
                full_text = card_text

            # Categorize based on label
            if "prompt" in label:
                prompt_cards.append([full_text])
                # Show preview with newlines replaced for display
                preview = full_text.replace('\n', '\\n')[:50]
                print(f"Row {row_idx}: Black card - {preview}...")
            elif "response" in label:
                response_cards.append([full_text])
                preview = full_text.replace('\n', '\\n')[:50]
                print(f"Row {row_idx}: White card - {preview}...")

    # Write prompt cards CSV
    if prompt_cards:
        with open(output_prompt, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Card Text', 'Tag1', 'Tag2', 'Tag3', 'Tag4', 'Tag5', 'Tag6', 'Tag7', 'Tag8', 'Tag9', 'Tag10'])
            writer.writerows(prompt_cards)
        print(f"Wrote {len(prompt_cards)} prompt cards to {output_prompt}")

    # Write response cards CSV
    if response_cards:
        with open(output_response, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Card Text', 'Tag1', 'Tag2', 'Tag3', 'Tag4', 'Tag5', 'Tag6', 'Tag7', 'Tag8', 'Tag9', 'Tag10'])
            writer.writerows(response_cards)
        print(f"Wrote {len(response_cards)} response cards to {output_response}")


def main():
    if len(sys.argv) != 4:
        print("Usage: python convert_spreadsheet_to_csv.py input_file.xlsx output_prompt.csv output_response.csv")
        sys.exit(1)

    input_file = Path(sys.argv[1])
    output_prompt = Path(sys.argv[2])
    output_response = Path(sys.argv[3])

    if not input_file.exists():
        print(f"Error: Input file not found: {input_file}")
        sys.exit(1)

    # Determine file type and process accordingly
    if input_file.suffix.lower() == '.xlsx':
        process_xlsx(input_file, output_prompt, output_response)
    elif input_file.suffix.lower() == '.ods':
        process_ods(input_file, output_prompt, output_response)
    else:
        print(f"Error: Unsupported file format: {input_file.suffix}")
        print("Supported formats: .xlsx, .ods")
        sys.exit(1)

    print("\nConversion complete!")


if __name__ == "__main__":
    main()
